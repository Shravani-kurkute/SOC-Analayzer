import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LABEL_FONT = Font(bold=True, size=10)
VALUE_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _style_row(ws, row, cols, bold_first=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = LABEL_FONT if (bold_first and c == 1) else VALUE_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left" if c == 1 else "center")


def build_excel(report_data: dict, report_type: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.replace("_", " ").title()

    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"SentinelAI - {report_type.replace('_', ' ').title()} Report"
    title_cell.font = Font(size=14, bold=True, color="1a1a2e")
    title_cell.alignment = Alignment(horizontal="center")

    row = 3

    if report_type == "executive":
        ws.cell(row=row, column=1, value="Metric")
        ws.cell(row=row, column=2, value="Value")
        _style_header(ws, row, 2)
        row += 1
        for key in ("total_incidents", "critical_incidents", "resolved_incidents", "open_incidents",
                     "avg_response_time_seconds", "avg_resolution_time_seconds", "soc_health_score"):
            ws.cell(row=row, column=1, value=key.replace("_", " ").title())
            ws.cell(row=row, column=2, value=report_data.get(key, ""))
            _style_row(ws, row, 2, bold_first=True)
            row += 1

        for section, key_field, val_field in [
            ("Top Risks", "name", "count"),
            ("Top Attack Types", "name", "count"),
            ("Top Countries", "name", "count"),
        ]:
            items = report_data.get(section.lower().replace(" ", "_"), [])
            if items:
                row += 1
                ws.cell(row=row, column=1, value=section).font = Font(bold=True, size=11)
                row += 1
                ws.cell(row=row, column=1, value="Name")
                ws.cell(row=row, column=2, value="Count")
                _style_header(ws, row, 2)
                row += 1
                for item in items:
                    ws.cell(row=row, column=1, value=str(item.get(key_field, "")))
                    ws.cell(row=row, column=2, value=item.get(val_field, 0))
                    _style_row(ws, row, 2)
                    row += 1

    elif report_type == "threat":
        ws.cell(row=row, column=1, value="IOC Type")
        ws.cell(row=row, column=2, value="Count")
        _style_header(ws, row, 2)
        row += 1
        for r in report_data.get("ioc_summary", []):
            ws.cell(row=row, column=1, value=r.get("type", ""))
            ws.cell(row=row, column=2, value=r.get("count", 0))
            _style_row(ws, row, 2)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Attack Categories").font = Font(bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value="Category")
        ws.cell(row=row, column=2, value="Count")
        _style_header(ws, row, 2)
        row += 1
        for r in report_data.get("attack_categories", []):
            ws.cell(row=row, column=1, value=r.get("name", ""))
            ws.cell(row=row, column=2, value=r.get("count", 0))
            _style_row(ws, row, 2)
            row += 1

    elif report_type == "incident":
        ws.cell(row=row, column=1, value="Field")
        ws.cell(row=row, column=2, value="Value")
        _style_header(ws, row, 2)
        row += 1
        for key in ("incident_id", "incident_title", "severity", "status", "description"):
            ws.cell(row=row, column=1, value=key.replace("_", " ").title())
            ws.cell(row=row, column=2, value=str(report_data.get(key, "") or ""))
            _style_row(ws, row, 2, bold_first=True)
            row += 1

        if report_data.get("tasks"):
            row += 1
            ws.cell(row=row, column=1, value="Tasks").font = Font(bold=True, size=11)
            row += 1
            for c, h in enumerate(["Title", "Status", "Priority", "Assignee"], 1):
                ws.cell(row=row, column=c, value=h)
            _style_header(ws, row, 4)
            row += 1
            for t in report_data["tasks"]:
                ws.cell(row=row, column=1, value=t.get("title", ""))
                ws.cell(row=row, column=2, value=t.get("status", ""))
                ws.cell(row=row, column=3, value=t.get("priority", ""))
                ws.cell(row=row, column=4, value=t.get("assignee_name", ""))
                _style_row(ws, row, 4)
                row += 1

        if report_data.get("evidence"):
            row += 1
            ws.cell(row=row, column=1, value="Evidence").font = Font(bold=True, size=11)
            row += 1
            for c, h in enumerate(["Filename", "Type", "Size", "SHA256"], 1):
                ws.cell(row=row, column=c, value=h)
            _style_header(ws, row, 4)
            row += 1
            for e in report_data["evidence"]:
                ws.cell(row=row, column=1, value=e.get("filename", ""))
                ws.cell(row=row, column=2, value=e.get("file_type", ""))
                ws.cell(row=row, column=3, value=e.get("file_size", ""))
                ws.cell(row=row, column=4, value=e.get("sha256", ""))
                _style_row(ws, row, 4)
                row += 1

    elif report_type == "asset":
        for c, h in enumerate(["Asset ID", "Name", "Type", "Criticality", "Owner"], 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, 5)
        row += 1
        for a in report_data.get("assets", []):
            ws.cell(row=row, column=1, value=a.get("id", ""))
            ws.cell(row=row, column=2, value=a.get("name", ""))
            ws.cell(row=row, column=3, value=a.get("type", ""))
            ws.cell(row=row, column=4, value=a.get("criticality", ""))
            ws.cell(row=row, column=5, value=a.get("owner", ""))
            _style_row(ws, row, 5)
            row += 1

    elif report_type == "compliance":
        ws.cell(row=row, column=1, value="Framework")
        ws.cell(row=row, column=2, value="Status")
        ws.cell(row=row, column=3, value="Coverage %")
        _style_header(ws, row, 3)
        row += 1
        for framework in ("soc2", "iso27001", "nist", "cis"):
            cov = report_data.get(f"{framework}_coverage", {})
            ws.cell(row=row, column=1, value=framework.upper())
            ws.cell(row=row, column=2, value=cov.get("status", ""))
            ws.cell(row=row, column=3, value=cov.get("percentage", 0))
            _style_row(ws, row, 3, bold_first=True)
            row += 1

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
